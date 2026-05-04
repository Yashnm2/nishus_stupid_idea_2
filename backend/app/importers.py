from __future__ import annotations

import csv
import re
import zipfile
from datetime import date, datetime, time, timedelta
from io import BytesIO, StringIO
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook
from pypdf import PdfReader

from .models import ClassSession


DAYS = {
    "monday": 0,
    "mon": 0,
    "tuesday": 1,
    "tue": 1,
    "wednesday": 2,
    "wed": 2,
    "thursday": 3,
    "thu": 3,
    "friday": 4,
    "fri": 4,
    "saturday": 5,
    "sat": 5,
    "sunday": 6,
    "sun": 6,
}

ROW_RE = re.compile(
    r"(?P<day>mon(?:day)?|tue(?:sday)?|wed(?:nesday)?|thu(?:rsday)?|fri(?:day)?|sat(?:urday)?|sun(?:day)?)"
    r"[\s,]+(?P<start>\d{1,2}(?::\d{2})?\s*(?:am|pm)?)\s*(?:-|–|to)\s*"
    r"(?P<end>\d{1,2}(?::\d{2})?\s*(?:am|pm)?)\s+(?P<module>.+)",
    re.IGNORECASE,
)


def week_start(today: date | None = None) -> date:
    today = today or date.today()
    return today - timedelta(days=today.weekday())


def parse_clock(value: str) -> time:
    raw = value.strip().lower().replace(" ", "")
    formats = ["%H:%M", "%H", "%I:%M%p", "%I%p"]
    for fmt in formats:
        try:
            return datetime.strptime(raw, fmt).time()
        except ValueError:
            pass
    raise ValueError(f"Unsupported time value: {value}")


def parse_entry(row: str, source: str) -> ClassSession | None:
    clean = " ".join(row.split())
    match = ROW_RE.search(clean)
    if not match:
        return None
    day_name = match.group("day").lower()
    day_index = DAYS[day_name[:3]]
    start_time = parse_clock(match.group("start"))
    end_time = parse_clock(match.group("end"))
    class_date = week_start() + timedelta(days=day_index)
    start = datetime.combine(class_date, start_time)
    end = datetime.combine(class_date, end_time)
    if end <= start:
        end += timedelta(hours=12)
    module = match.group("module").strip(" -")
    return ClassSession(day=class_date.strftime("%A"), start=start, end=end, module=module, source=source)


def _sessions_from_rows(rows: list[str], source: str) -> list[ClassSession]:
    sessions = []
    for row in rows:
        session = parse_entry(row, source)
        if session:
            sessions.append(session)
    return sessions


def parse_csv(data: bytes, filename: str) -> list[ClassSession]:
    text = data.decode("utf-8-sig")
    rows = []
    for row in csv.reader(StringIO(text)):
        rows.append(" ".join(cell for cell in row if cell))
    return _sessions_from_rows(rows, filename)


def parse_xlsx(data: bytes, filename: str) -> list[ClassSession]:
    workbook = load_workbook(BytesIO(data), data_only=True)
    rows = []
    for sheet in workbook.worksheets:
        for values in sheet.iter_rows(values_only=True):
            rows.append(" ".join(str(value) for value in values if value is not None))
    return _sessions_from_rows(rows, filename)


def parse_pdf(data: bytes, filename: str) -> list[ClassSession]:
    reader = PdfReader(BytesIO(data))
    rows = []
    for page in reader.pages:
        rows.extend((page.extract_text() or "").splitlines())
    return _sessions_from_rows(rows, filename)


def parse_ods(data: bytes, filename: str) -> list[ClassSession]:
    rows = []
    with zipfile.ZipFile(BytesIO(data)) as archive:
        xml = archive.read("content.xml")
    root = ElementTree.fromstring(xml)
    for row in root.iter():
        if row.tag.endswith("table-row"):
            values = [node.text.strip() for node in row.iter() if node.text and node.text.strip()]
            if values:
                rows.append(" ".join(values))
    return _sessions_from_rows(rows, filename)


def parse_timetable(filename: str, data: bytes) -> list[ClassSession]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        return parse_csv(data, filename)
    if suffix == ".xlsx":
        return parse_xlsx(data, filename)
    if suffix == ".pdf":
        return parse_pdf(data, filename)
    if suffix == ".ods":
        return parse_ods(data, filename)
    raise ValueError("Supported timetable formats are CSV, XLSX, ODS, and PDF.")
